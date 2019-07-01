import math
import numpy as np
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.model_selection import KFold, StratifiedKFold
from sklearn import metrics
from bayes_classifier import bayes_rule
from utils import ClassifierSelector, dataset_loader, plot_hist, plot_error
import matplotlib.pyplot as plt


def class_balance_test(gauss_distr):
    """
    :param gauss_distr: is a list that contains param of two 1-d gaussian distributions
           [mean1, deviation1, mean2, deviation2]
    :return: print results in a excel file
    """
    wb = load_workbook("error-estimates.xlsx")
    sheet1 = wb['Foglio1']
    row = 4
    column = 2
    test_list = [[100, 100], [100, 200], [200, 100], [200, 200], [200, 400], [400, 200], [500, 500], [700, 1000], [1000, 700], [1000, 1000]]
    mu1 = gauss_distr[0]
    sigma1 = gauss_distr[1]
    mu2 = gauss_distr[2]
    sigma2 = gauss_distr[3]
    x1 = np.random.normal(mu1, sigma1, 1000)
    x2 = np.random.normal(mu2, sigma2, 1000)
    for test in test_list:
        x = np.concatenate((x1[:test[0]], x2[:test[1]]), axis=0)
        y1 = np.zeros(test[0])
        y2 = np.full(test[1], 1)
        y = np.concatenate((y1, y2), axis=0)

        p1 = test[0]/(test[0]+test[1])  # prior probability
        p2 = test[1] / (test[0] + test[1])
        y_pred, e1, e2, b1, b2 = bayes_rule(x, mu1, sigma1, mu2, sigma2, p1, p2)

        conf_matrix = metrics.confusion_matrix(y, y_pred)
        tmp1 = conf_matrix[0, 1] / list(y).count(0)
        tmp2 = conf_matrix[1, 0] / list(y).count(1)
        tmp = (conf_matrix[0, 1] + conf_matrix[1, 0]) / len(y)
        sheet1.cell(row=row, column=column).value = e1
        sheet1.cell(row=row + 1, column=column).value = e2
        sheet1.cell(row=row + 2, column=column).value = (e1+e2)
        sheet1.cell(row=row + 3, column=column).value = tmp1
        sheet1.cell(row=row + 4, column=column).value = tmp2
        sheet1.cell(row=row + 5, column=column).value = tmp
        wb.save("error-estimates.xlsx")
        column += 1


def dataset_test(classifier, validation, sample_estimate=False, shuffle=True, real_dataset=False):
    """
    :param classifier: choose between bayes, kNN, MLP and Tree
    :param validation: choose between resub, holdout and cross
    :param sample_estimate: works only with bayes classifier
    :param shuffle: if False dataset is composed by all class1 elements then all class2 elements
    :param real_dataset: if True use the bank loan dataset, else dataset is generated from two 1-d gaussian distribution
           does not work with bayes classifier
    :return: print results in a excel file
    """
    wb = load_workbook("error-estimates.xlsx")
    sheet1 = wb['Foglio1']
    row = 29
    column = 3
    error1 = []
    error2 = []
    error = []
    test_list = [100, 200, 500, 1000, 2000, 5000, 10000, 20000, 50000]
    e1, e2, b1, b2, tmp1, tmp2, tmp = 0, 0, 0, 0, 0, 0, 0
    mu1 = 0
    sigma1 = math.sqrt(1)
    mu2 = 0
    sigma2 = math.sqrt(0.25)
    clf = ClassifierSelector(classifier)

    for test in test_list:
        for i in range(10):
            if real_dataset:
                x, y = dataset_loader(test)
                if shuffle:
                    shuffle_idx = np.arange(len(y))
                    np.random.shuffle(shuffle_idx)
                    x = x[shuffle_idx, :]
                    y = y[shuffle_idx]
            else:
                x1 = np.random.normal(mu1, sigma1, test)
                x2 = np.random.normal(mu2, sigma2, test)
                x = np.concatenate((x1, x2), axis=0)
                y1 = np.zeros(test)
                y2 = np.full(test, 1)
                y = np.concatenate((y1, y2), axis=0)
                if shuffle:
                    shuffle_idx = np.arange(len(y))
                    np.random.shuffle(shuffle_idx)
                    x = x[shuffle_idx]
                    y = y[shuffle_idx]

            if validation == 'resub':
                if classifier == 'bayes':
                    if sample_estimate:
                        mu1 = np.mean(x1)
                        mu2 = np.mean(x2)
                        sigma1 = math.sqrt(np.var(x1))
                        sigma2 = math.sqrt(np.var(x2))
                    y_pred, e1, e2, b1, b2 = bayes_rule(x, mu1, sigma1, mu2, sigma2, 0.5, 0.5)
                else:
                    clf.fit(x, y)
                    y_pred = clf.predict(x)
                conf_matrix = metrics.confusion_matrix(y, y_pred)
                tmp1 = conf_matrix[0, 1] / list(y).count(0)
                tmp2 = conf_matrix[1, 0] / list(y).count(1)
                tmp = (conf_matrix[0, 1] + conf_matrix[1, 0]) / len(y)

            if validation == 'holdout':
                x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.4, random_state=0, stratify=y)

                if classifier == 'bayes':
                    if sample_estimate:
                        x1 = x_train[y_train == 0]
                        x2 = x_train[y_train == 1]
                        mu1 = np.mean(x1)
                        mu2 = np.mean(x2)
                        sigma1 = math.sqrt(np.var(x1))
                        sigma2 = math.sqrt(np.var(x2))
                    y_pred, e1, e2, b1, b2 = bayes_rule(x_test, mu1, sigma1, mu2, sigma2, 0.5, 0.5)
                else:
                    clf.fit(x_train, y_train)
                    y_pred = clf.predict(x_test)

                conf_matrix = metrics.confusion_matrix(y_test, y_pred)
                tmp1 = conf_matrix[0, 1] / list(y_test).count(0)
                tmp2 = conf_matrix[1, 0] / list(y_test).count(1)
                tmp = (conf_matrix[0, 1] + conf_matrix[1, 0]) / len(y_test)

            if validation == 'cross':
                cross1 = []
                cross2 = []
                cross = []
                if len(x.shape) == 1:
                    x = np.reshape(x, [len(x), 1])
                skf = StratifiedKFold(n_splits=10)
                skf.get_n_splits(x, y)

                for train_index, test_index in skf.split(x, y):
                    x_train, x_test = x[train_index], x[test_index]
                    y_train, y_test = y[train_index], y[test_index]

                    if classifier == 'bayes':
                        if sample_estimate:
                            x1 = x_train[y_train == 0]
                            x2 = x_train[y_train == 1]
                            mu1 = np.mean(x1)
                            mu2 = np.mean(x2)
                            sigma1 = math.sqrt(np.var(x1))
                            sigma2 = math.sqrt(np.var(x2))
                        y_pred, e1, e2, b1, b2 = bayes_rule(x_test, mu1, sigma1, mu2, sigma2, 0.5, 0.5)
                    else:
                        clf.fit(x_train, y_train)
                        y_pred = clf.predict(x_test)

                    conf_matrix = metrics.confusion_matrix(y_test, y_pred)
                    c1 = conf_matrix[0, 1] / list(y_test).count(0)
                    c2 = conf_matrix[1, 0] / list(y_test).count(1)
                    c = (conf_matrix[0, 1] + conf_matrix[1, 0]) / len(y_test)
                    cross1.append(c1)
                    cross2.append(c2)
                    cross.append(c)

                tmp1 = np.average(cross1)
                tmp2 = np.average(cross2)
                tmp = np.average(cross)
            error1.append(tmp1)
            error2.append(tmp2)
            error.append(tmp)

        sheet1.cell(row=row, column=column).value = np.average(error1)
        sheet1.cell(row=row + 1, column=column).value = np.var(error1)
        sheet1.cell(row=row + 2, column=column).value = np.average(error2)
        sheet1.cell(row=row + 3, column=column).value = np.var(error2)
        sheet1.cell(row=row + 4, column=column).value = np.average(error)
        sheet1.cell(row=row + 5, column=column).value = np.var(error)

        if test == 100:
            row1 = 14
            col1 = 2
            for er in error1:
                sheet1.cell(row=row1, column=col1).value = er
                row1 += 1
            row1 = 14
            col1 = 3
            for er in error2:
                sheet1.cell(row=row1, column=col1).value = er
                row1 += 1
            row1 = 14
            col1 = 4
            for er in error:
                sheet1.cell(row=row1, column=col1).value = er
                row1 += 1

        wb.save("error-estimates.xlsx")
        column += 1
        info = [classifier, validation, test]
        # plot_hist(x1, x2)
        plot_error(error, e1+e2, info)
        error = []
    print("Bayes error1", e1)
    print("Bayes error2", e2)
    print("Bayes error", e1+e2)
    print("Bayes border: {}, {}".format(b1, b2))
